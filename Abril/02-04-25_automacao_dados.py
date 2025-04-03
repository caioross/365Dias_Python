import openpyxl

def criar_planilha(nome_arquivo):
    """
    Função para criar uma nova planilha Excel e adicionar algumas informações.
    
    :param nome_arquivo: Nome do arquivo Excel a ser criado.
    """
    try:
        # Cria uma nova planilha
        wb = openpyxl.Workbook()
        
        # Acessa a planilha ativa (por padrão, a planilha é chamada "Sheet")
        sheet = wb.active
        sheet.title = "Dados Exemplo"
        
        # Adiciona cabeçalhos nas células
        sheet['A1'] = "Nome"
        sheet['B1'] = "Idade"
        sheet['C1'] = "Cidade"
        
        # Adiciona algumas linhas de dados
        dados = [
            ['João', 25, 'São Paulo'],
            ['Maria', 30, 'Rio de Janeiro'],
            ['Pedro', 28, 'Belo Horizonte']
        ]
        
        for i, linha in enumerate(dados, start=2):
            sheet[f'A{i}'] = linha[0]
            sheet[f'B{i}'] = linha[1]
            sheet[f'C{i}'] = linha[2]
        
        # Salva o arquivo Excel
        wb.save(nome_arquivo)
        print(f"Planilha criada com sucesso: {nome_arquivo}")
    
    except Exception as e:
        print(f"Erro ao criar a planilha: {e}")

def ler_planilha(nome_arquivo):
    """
    Função para ler e exibir os dados de uma planilha Excel.
    
    :param nome_arquivo: Caminho do arquivo Excel a ser lido.
    """
    try:
        # Abre a planilha existente
        wb = openpyxl.load_workbook(nome_arquivo)
        
        # Acessa a primeira planilha
        sheet = wb.active
        
        # Itera sobre as células e imprime os valores
        print(f"Conteúdo da planilha {nome_arquivo}:")
        for row in sheet.iter_rows(values_only=True):
            print(row)
    
    except Exception as e:
        print(f"Erro ao ler a planilha: {e}")

def main():
    """
    Função principal para criar e ler uma planilha Excel.
    """
    nome_arquivo = input("Digite o nome do arquivo Excel (com extensão .xlsx): ")
    
    # Cria a planilha
    criar_planilha(nome_arquivo)
    
    # Lê a planilha e exibe os dados
    ler_planilha(nome_arquivo)

if __name__ == "__main__":
    main()
